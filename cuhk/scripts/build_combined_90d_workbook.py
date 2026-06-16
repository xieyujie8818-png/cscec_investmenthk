# -*- coding: utf-8 -*-
"""Build single-sheet Excel: last-90-day rent + sale for 14 estates (Midland, Centaline fallback)."""
from __future__ import annotations

import json
import re
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "周邊樓盤租賃及成交對標表.xlsx"
OUT_CSV = ROOT / "周邊樓盤租賃及成交對標表.csv"
CACHE = ROOT / "scripts" / "estate_90d_cache.json"

AS_OF = date(2026, 6, 10)
CUTOFF = AS_OF - timedelta(days=90)  # 2026-03-12
HEADERS_HTTP = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

ESTATES = [
    {
        "seq": 1,
        "name": "大埔寶馬山",
        "en": "Pacific Palisades",
        "dev": "信和",
        "loc": "山賢路8號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E5%A4%A7%E5%9F%94%E5%AF%B6%E9%A6%AC%E5%B1%B1-E11504",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E5%A4%A7%E5%9F%94%E5%AF%B6%E9%A6%AC%E5%B1%B1_2-DCRVURVXRO",
    },
    {
        "seq": 2,
        "name": "雍怡雅苑",
        "en": "Chateau Royale",
        "dev": "新鴻基",
        "loc": "雍宜路1號",
        "midland_url": "https://www.midland.com.hk/zh-hk/property/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E9%9B%8D%E6%80%A1%E9%9B%85%E8%8B%91%E9%9B%8D%E6%97%A5%E5%BA%AD%E9%9B%8D%E5%AE%9C%E8%B7%AF1%E8%99%9F-%E7%8D%A8%E7%AB%8B%E5%B1%8B-M300300767",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E9%9B%8D%E6%80%A1%E9%9B%85%E8%8B%91_2-DCRVURVXRO",
    },
    {
        "seq": 3,
        "name": "泓山",
        "en": "Hampstead",
        "dev": "建灝",
        "loc": "逸遙路3號",
        "midland_url": None,
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E6%B3%93%E5%B1%B1_2-DCRVURVXRO",
    },
    {
        "seq": 4,
        "name": "天鑽",
        "en": "The Regent",
        "dev": "新鴻基",
        "loc": "山塘路8號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E5%A4%A9%E9%91%BD-E000016716",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E5%A4%A9%E9%91%BD_2-DEPPWPPJPB",
    },
    {
        "seq": 5,
        "name": "悠然山莊",
        "en": "The Paragon",
        "dev": "加文",
        "loc": "山賢路9號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E6%82%A0%E7%84%B6%E5%B1%B1%E8%8E%8A-E11584",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E6%82%A0%E7%84%B6%E5%B1%B1%E8%8E%8A_2-DCRVURVXRO",
    },
    {
        "seq": 6,
        "name": "偉東·雍宜山莊",
        "en": "Grand Dynasty View",
        "dev": "—",
        "loc": "下黃宜坳",
        "midland_url": None,
        "centanet_url": None,
    },
    {
        "seq": 7,
        "name": "疊翠豪庭",
        "en": "The Paramount (4188)",
        "dev": "—",
        "loc": "大埔公路大埔滘段4188號",
        "midland_url": None,
        "centanet_url": None,
    },
    {
        "seq": 8,
        "name": "峰林軒",
        "en": "Daisyfield",
        "dev": "—",
        "loc": "大埔公路大埔滘段4135號",
        "midland_url": None,
        "centanet_url": None,
    },
    {
        "seq": 9,
        "name": "史提福樓",
        "en": "Stafford House",
        "dev": "—",
        "loc": "大埔滘新圍",
        "midland_url": None,
        "centanet_url": None,
    },
    {
        "seq": 10,
        "name": "逍遙雋岸",
        "en": "The Gables",
        "dev": "興聯置業",
        "loc": "逸遙路大埔滘段18號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E9%80%8D%E9%81%99%E9%9B%8B%E5%B2%B8-E000004058",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E9%80%8D%E9%81%99%E9%9B%8B%E5%B2%B8_2-DCRVURVXRO",
    },
    {
        "seq": 11,
        "name": "翡翠花園",
        "en": "Savanna Garden",
        "dev": "新鴻基",
        "loc": "大埔公路大埔滘段4283號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E7%BF%A1%E7%BF%A0%E8%8A%B1%E5%9C%92-E00070",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E7%BF%A1%E7%BF%A0%E8%8A%B1%E5%9C%92_2-DCRVURVXRO",
    },
    {
        "seq": 12,
        "name": "盈峰翠邸",
        "en": "The Paramount",
        "dev": "長江實業",
        "loc": "山塘路23號",
        "midland_url": "https://www.midland.com.hk/zh-hk/property/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E7%9B%88%E5%B3%B0%E7%BF%A0%E9%82%B802%E5%BA%A7%E4%B8%AD%E5%B1%A4B%E5%AE%A4-M300551705",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E7%9B%88%E5%B3%B0%E7%BF%A0%E9%82%B8_2-DCRVURVXRO",
    },
    {
        "seq": 13,
        "name": "雲滙",
        "en": "St Martin",
        "dev": "新鴻基",
        "loc": "白石角科進路12號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E7%99%BD%E7%9F%B3%E8%A7%92-%E9%9B%B2%E6%BB%99-E000016471",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E9%9B%B2%E6%BB%99_3-DESPWPPHPW",
    },
    {
        "seq": 14,
        "name": "海日灣II",
        "en": "Centra Horizon",
        "dev": "億京發展",
        "loc": "白石角創新路18號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E7%99%BD%E7%9F%B3%E8%A7%92-%E6%B5%B7%E6%97%A5%E7%81%A3II-E000016718",
        "centanet_url": "https://hk.centanet.com/findproperty/list/transaction/%E6%B5%B7%E6%97%A5%E7%81%A3II_2-DEPPWPPJPB",
    },
]

# Verified Midland transactions within 90d (2026-03-12 – 2026-06-10), used when fetch returns empty shell.
VERIFIED_SALES = {
    "大埔寶馬山": [
        (728, 10440, "26/04/26", False),
        (533, 10038, "16/04/26", False),
        (756, 11376, "16/04/26", False),
        (565, 10938, "07/05/26", False),
        (996, 10442, "18/05/26", False),
    ],
    "悠然山莊": [
        (913, 9069, "01/04/26", False),
        (914, 9497, "16/04/26", False),
    ],
    "天鑽": [
        (457, 13239, "30/04/26", False),
        (735, 13306, "28/04/26", False),
        (457, 13961, "24/04/26", False),
        (525, 12533, "21/04/26", False),
        (457, 15011, "21/04/26", False),
        (524, 13359, "20/04/26", False),
        (551, 13721, "27/05/26", False),
        (584, 13116, "26/05/26", False),
        (775, 12490, "23/05/26", False),
        (524, 13130, "21/05/26", False),
        (748, 12687, "17/05/26", False),
        (551, 14338, "09/06/26", False),
        (744, 11559, "09/06/26", False),
        (735, 12109, "07/06/26", False),
        (509, 12770, "03/06/26", False),
        (529, 13384, "03/06/26", False),
    ],
    "盈峰翠邸": [(916, 9279, "28/05/26", False)],
    "雲滙": [
        (629, 13800, "09/06/26", False),
        (361, 11911, "09/06/26", False),
        (714, 15238, "08/06/26", False),
        (442, 14593, "03/06/26", False),
        (451, 14634, "28/05/26", False),
        (278, 15647, "28/05/26", False),
        (379, 15198, "27/05/26", False),
        (510, 14902, "25/05/26", False),
        (501, 15305, "21/05/26", False),
        (457, 14486, "19/05/26", False),
    ],
    "海日灣II": [
        (608, 13306, "09/06/26", False),
        (1058, 13948, "08/06/26", False),
        (1039, 13936, "07/06/26", False),
        (1019, 14228, "07/06/26", False),
        (629, 13180, "02/06/26", False),
        (304, 14408, "02/06/26", False),
        (1022, 15268, "01/06/26", False),
        (1038, 15830, "01/06/26", False),
    ],
    "翡翠花園": [
        (959, 9518, "12/03/26", False),
        (1987, 13840, "26/03/26", False),
    ],
}

VERIFIED_RENT_TX = {
    "天鑽": [
        (617, 20000, 38, "21/04/26"),
        (372, 20000, 43, "01/05/26"),
        (524, 20000, 37, "01/05/26"),
        (524, 20000, 40, "11/05/26"),
        (593, 30000, 47, "11/05/26"),
    ],
}

VERIFIED_LISTING = {
    "大埔寶馬山": {"rent_min": 18000, "rent_max": 30000, "rent_psf": 32, "sale_psf": 11304},
    "悠然山莊": {"rent_min": 36000, "rent_max": 36000, "rent_psf": 29, "sale_psf": 10847},
    "天鑽": {"rent_min": 15000, "rent_max": 31000, "rent_psf": 41, "sale_psf": 14420},
    "翡翠花園": {"rent_min": 28500, "rent_max": 40000, "rent_psf": 33, "sale_psf": 10897},
    "雲滙": {"rent_min": 13000, "rent_max": 59000, "rent_psf": 43, "sale_psf": 16003},
    "海日灣II": {"rent_min": 12300, "rent_max": 50000, "rent_psf": 46, "sale_psf": 15833},
}


def fetch(url: str) -> str:
    req = urllib.request.Request(url, headers=HEADERS_HTTP)
    with urllib.request.urlopen(req, timeout=45) as r:
        return r.read().decode("utf-8", "replace")


def parse_date(d: str) -> date | None:
    m = re.match(r"(\d{2})/(\d{2})/(\d{2})", d)
    if not m:
        return None
    dd, mm, yy = map(int, m.groups())
    return date(2000 + yy, mm, dd)


def in_window(d: str) -> bool:
    dt = parse_date(d)
    return dt is not None and CUTOFF <= dt <= AS_OF


def parse_midland_sales(html: str) -> list[dict]:
    out = []
    for m in re.finditer(
        r"(\d{2}/\d{2}/26)[^$]{0,350}?實(\d+)呎[^$]{0,150}?@\$([\d,]+)",
        html,
    ):
        chunk = m.group(0)
        psf = int(m.group(3).replace(",", ""))
        if psf < 5000:
            continue
        special = "^" in chunk or "車位" in chunk[:120]
        out.append(
            {
                "date": m.group(1),
                "area": int(m.group(2)),
                "psf": psf,
                "special": special,
            }
        )
    # dedupe by date+area+psf
    seen = set()
    deduped = []
    for x in out:
        k = (x["date"], x["area"], x["psf"])
        if k not in seen:
            seen.add(k)
            deduped.append(x)
    return deduped


def parse_midland_rent(html: str) -> list[dict]:
    out = []
    for m in re.finditer(
        r"(\d{2}/\d{2}/26)[^$]{0,350}?實(\d+)呎[^$]{0,120}?\$([\d,]+(?:\.\d+)?)萬[^$]{0,40}?@\$([\d,]+)",
        html,
    ):
        psf = int(m.group(4).replace(",", ""))
        if psf >= 200:
            continue
        out.append(
            {
                "date": m.group(1),
                "area": int(m.group(2)),
                "rent": int(float(m.group(3).replace(",", "")) * 10000),
                "psf": psf,
            }
        )
    for m in re.finditer(
        r"(\d{2}/\d{2}/26)[^$]{0,350}?實(\d+)呎[^$]{0,120}?租\s*\$([\d,]+)[^$]{0,40}?@\$([\d,]+)",
        html,
    ):
        psf = int(m.group(4).replace(",", ""))
        out.append(
            {
                "date": m.group(1),
                "area": int(m.group(2)),
                "rent": int(m.group(3).replace(",", "")),
                "psf": psf,
            }
        )
    seen = set()
    deduped = []
    for x in out:
        k = (x["date"], x["area"], x["rent"], x["psf"])
        if k not in seen:
            seen.add(k)
            deduped.append(x)
    return deduped


def extract_listing_summary(html: str) -> dict:
    s = {}
    m = re.search(r"放盤平均呎租\s*\$([\d,]+)", html)
    if m:
        s["rent_psf"] = int(m.group(1).replace(",", ""))
    m = re.search(r"放盤租價\s*\$([\d,]+)\s*-\s*\$([\d,]+)", html)
    if m:
        s["rent_min"] = int(m.group(1).replace(",", ""))
        s["rent_max"] = int(m.group(2).replace(",", ""))
    m = re.search(r"放盤平均呎價\s*\$([\d,]+)", html)
    if m:
        s["sale_psf"] = int(m.group(1).replace(",", ""))
    m = re.search(r"成交平均呎租 \(過去30日\)\s*\$([\d,]+)", html)
    if m:
        s["rent_tx_30d"] = int(m.group(1).replace(",", ""))
    return s


def summarize_sales(deals: list[tuple], exclude_special: bool = True) -> dict | None:
    sub = [d for d in deals if in_window(d[2]) and (not exclude_special or not d[3])]
    if not sub:
        return None
    psfs = [d[1] for d in sub]
    total = sum(d[0] * d[1] for d in sub)
    area_sum = sum(d[0] for d in sub)
    return {
        "count": len(sub),
        "psf_min": min(psfs),
        "psf_max": max(psfs),
        "psf_avg": round(total / area_sum),
    }


def summarize_rent_tx(rows: list[tuple]) -> dict | None:
    sub = [r for r in rows if in_window(r[3])]
    if not sub:
        return None
    rents = [r[1] for r in sub]
    psfs = [r[2] for r in sub]
    return {
        "count": len(sub),
        "rent_min": min(rents),
        "rent_max": max(rents),
        "psf_min": min(psfs),
        "psf_max": max(psfs),
        "dtype": "租務成交",
    }


def summarize_rent_listing(lst: dict) -> dict | None:
    if not lst.get("rent_min"):
        return None
    return {
        "rent_min": lst["rent_min"],
        "rent_max": lst["rent_max"],
        "psf_min": lst.get("rent_psf"),
        "psf_max": lst.get("rent_psf"),
        "dtype": "屋苑放盤參考",
    }


def fmt_range(lo, hi, suffix=""):
    if lo is None or hi is None:
        return "—"
    if lo == hi:
        return f"{lo:,}{suffix}"
    return f"{lo:,} – {hi:,}{suffix}"


def collect_estate(est: dict) -> dict:
    name = est["name"]
    result = {
        "source": "—",
        "sale_count": 0,
        "sale_psf_rng": "—",
        "sale_psf_avg": "—",
        "rent_rng": "—",
        "rent_psf_rng": "—",
        "rent_dtype": "—",
        "note": "",
    }

    sales_raw: list[tuple] = []
    rent_raw: list[tuple] = []
    listing = VERIFIED_LISTING.get(name, {})

    if est.get("midland_url"):
        try:
            html = fetch(est["midland_url"])
            if name in html or len(html) > 5000:
                result["source"] = "美聯物業"
                for s in parse_midland_sales(html):
                    sales_raw.append((s["area"], s["psf"], s["date"], s["special"]))
                for r in parse_midland_rent(html):
                    rent_raw.append((r["area"], r["rent"], r["psf"], r["date"]))
                listing = {**listing, **extract_listing_summary(html)}
        except Exception as e:
            result["note"] += f"美聯抓取失敗({e}); "

    # fallback verified when parser got nothing
    if not sales_raw and name in VERIFIED_SALES:
        sales_raw = VERIFIED_SALES[name]
        if result["source"] == "—":
            result["source"] = "美聯物業（人工核實）"

    if not rent_raw and name in VERIFIED_RENT_TX:
        rent_raw = [
            (a, rent, psf, d) for a, rent, psf, d in VERIFIED_RENT_TX[name]
        ]

    sale_sum = summarize_sales(sales_raw)
    if sale_sum:
        result["sale_count"] = sale_sum["count"]
        result["sale_psf_rng"] = fmt_range(sale_sum["psf_min"], sale_sum["psf_max"])
        result["sale_psf_avg"] = f"{sale_sum['psf_avg']:,}"
    else:
        result["note"] += "近90日無二手成交; "

    rent_sum = summarize_rent_tx(rent_raw)
    if not rent_sum:
        rent_sum = summarize_rent_listing(listing)
    if rent_sum:
        result["rent_rng"] = fmt_range(rent_sum["rent_min"], rent_sum["rent_max"])
        psf_lo = rent_sum.get("psf_min")
        psf_hi = rent_sum.get("psf_max")
        if psf_lo and psf_hi:
            result["rent_psf_rng"] = fmt_range(psf_lo, psf_hi)
        elif listing.get("rent_psf"):
            result["rent_psf_rng"] = f"{listing['rent_psf']:,}"
        result["rent_dtype"] = rent_sum.get("dtype", "租務成交")
    else:
        if est.get("centanet_url") and result["source"] == "—":
            result["source"] = "中原地產（待查）"
            result["note"] += "美聯無租務數據，建議查中原; "
        else:
            result["note"] += "近90日無租務成交／放盤參考; "

    urls = []
    if est.get("midland_url"):
        urls.append(f"美聯：{est['midland_url']}")
    if est.get("centanet_url"):
        urls.append(f"中原：{est['centanet_url']}")
    result["source"] = "；".join(urls) if urls else "—"

    parts = [p.strip() for p in re.split(r"[;；]+", result["note"]) if p.strip()]
    if sale_sum is None and est.get("centanet_url") and not any("中原" in p for p in parts):
        parts.append("成交可參考中原連結")
    result["note"] = "；".join(parts)
    return result


def build_workbook(rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "近90日租售對標"

    meta = [
        "大埔黃宜坳周邊樓盤 · 近90日租賃及二手成交對標（14個可比屋苑）",
        f"統計區間：{CUTOFF.isoformat()} 至 {AS_OF.isoformat()}（共90日）｜整理日期：{AS_OF.isoformat()}",
        "資料來源：美聯物業屋苑／放盤頁註冊處紀錄；美聯無資料時標註中原查詢連結。平均呎價為面積加權平均。",
        "註：疊翠豪庭（4188號）與盈峰翠邸（山塘路23號）英文名同為 The Paramount，為不同項目。",
    ]
    meta_fill = PatternFill("solid", fgColor="D9E2F3")
    for i, line in enumerate(meta, 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=14)
        c = ws.cell(row=i, column=1, value=line)
        c.fill = meta_fill
        c.alignment = Alignment(wrap_text=True)

    headers = [
        "序號",
        "項目名稱",
        "英文名",
        "發展商",
        "位置",
        "近90日月租範圍（元）",
        "近90月呎租範圍（元/呎）",
        "租賃資料類型",
        "近90日成交量（宗）",
        "近90日呎價範圍（元/呎）",
        "近90日平均呎價（元/呎）",
        "備註",
        "資料來源",
    ]
    start = len(meta) + 2
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=j, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for est, data in zip(ESTATES, rows):
        ws.append(
            [
                est["seq"],
                est["name"],
                est["en"],
                est["dev"],
                est["loc"],
                data["rent_rng"],
                data["rent_psf_rng"],
                data["rent_dtype"],
                data["sale_count"],
                data["sale_psf_rng"],
                data["sale_psf_avg"],
                data["note"],
                data["source"],
            ]
        )

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(min(len(str(c.value or "")) + 2, 50) for c in col)
        ws.column_dimensions[letter].width = max(width, 10)

    wb.save(OUT)
    print("Wrote", OUT)


def build_csv(rows: list[dict]):
    import csv

    fields = [
        "序號",
        "項目名稱",
        "英文名",
        "發展商",
        "位置",
        "近90日月租範圍_元",
        "近90月呎租範圍_元每呎",
        "租賃資料類型",
        "近90日成交量_宗",
        "近90日呎價範圍_元每呎",
        "近90日平均呎價_元每呎",
        "備註",
        "資料來源",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for est, data in zip(ESTATES, rows):
            w.writerow(
                {
                    "序號": est["seq"],
                    "項目名稱": est["name"],
                    "英文名": est["en"],
                    "發展商": est["dev"],
                    "位置": est["loc"],
                    "近90日月租範圍_元": data["rent_rng"],
                    "近90月呎租範圍_元每呎": data["rent_psf_rng"],
                    "租賃資料類型": data["rent_dtype"],
                    "近90日成交量_宗": data["sale_count"],
                    "近90日呎價範圍_元每呎": data["sale_psf_rng"],
                    "近90日平均呎價_元每呎": data["sale_psf_avg"],
                    "備註": data["note"],
                    "資料來源": data["source"],
                }
            )
    print("Wrote", OUT_CSV)


def main():
    rows = []
    cache = {}
    for est in ESTATES:
        data = collect_estate(est)
        rows.append(data)
        cache[est["name"]] = data
        print(
            est["seq"],
            "sale",
            data["sale_count"],
            "rent",
            data["rent_rng"],
        )

    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    build_workbook(rows)
    build_csv(rows)


if __name__ == "__main__":
    main()
