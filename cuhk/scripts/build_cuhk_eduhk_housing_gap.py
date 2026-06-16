# -*- coding: utf-8 -*-
"""Build CUHK + EdUHK off-campus housing gap table (Colliers formula)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "學生宿舍" / "中大教大校外住宿缺口测算.xlsx"

# ── 数据源 ──
# CUHK: Facts 2024（政府资助，2024/9/30）
CUHK_UG = {"total": 18438, "local": 15742, "nl": 2696, "beds": 9500, "beds_note": "九书院 + I-House"}
CUHK_PG = {"total": 4650, "local": 1699, "nl": 2951, "beds": 1100, "beds_note": "PGH 研宿"}

# EdUHK: 立法会 CB(3)315/2025 Table 2（2024/10/31，FTE）；宿位：教大官网 ~2,200
EDUHK_UG = {"total": 5336, "local": 4737, "nl": 599, "beds": 2200, "beds_note": "四座学生宿舍"}
EDUHK_PG = {"total": 601, "local": 447, "nl": 154, "beds": 0, "beds_note": "与本科生共用，无独立研宿"}


def colliers_eff(local_ft: int, nl: int) -> int:
    return round(local_ft * 0.35 + nl)


def metrics(row: dict) -> dict:
    eff = colliers_eff(row["local"], row["nl"])
    gap = max(0, eff - row["beds"])
    cov = f"{row['beds'] / row['total'] * 100:.0f}%" if row["beds"] and row["total"] else "—"
    return {"eff": eff, "gap": gap, "cov": cov}


def school_summary(ug: dict, pg: dict) -> dict:
    m_ug, m_pg = metrics(ug), metrics(pg)
    total_eff = m_ug["eff"] + m_pg["eff"]
    total_beds = ug["beds"] + pg["beds"]
    total_gap = m_ug["gap"] + m_pg["gap"]
    all_local = ug["local"] + pg["local"]
    all_nl = ug["nl"] + pg["nl"]
    all_total = ug["total"] + pg["total"]
    return {
        "total": all_total,
        "local": all_local,
        "nl": all_nl,
        "eff": total_eff,
        "beds": total_beds,
        "gap": total_gap,
        "ug": m_ug,
        "pg": m_pg,
    }


CUHK = school_summary(CUHK_UG, CUHK_PG)
EDUHK = school_summary(EDUHK_UG, EDUHK_PG)

SENTENCE_1 = (
    f"2024/25学年，香港中文大学政府资助全日制在校生约{CUHK['total']/10000:.2f}万人"
    f"（本科{CUHK_UG['total']/10000:.2f}万、研究生{CUHK_PG['total']/10000:.2f}万），"
    f"其中本地约{CUHK['local']/10000:.2f}万、非本地约{CUHK['nl']/10000:.2f}万；"
    f"香港教育大学政府资助在校生约{EDUHK['total']/10000:.2f}万人"
    f"（本科{EDUHK_UG['total']/10000:.2f}万、研究生{EDUHK_PG['total']/10000:.2f}万），"
    f"其中本地约{EDUHK['local']/10000:.2f}万、非本地约{EDUHK['nl']/10000:.2f}万。"
)

SENTENCE_2 = (
    "参照高力国际测算逻辑（须住宿人数＝本地全日制在校生×35%＋非本地在校生×100%），"
    f"中大有效住宿需求约{CUHK['eff']/10000:.2f}万床、校内供给约{CUHK['beds']/10000:.2f}万床、"
    f"校外缺口约{CUHK['gap']/10000:.2f}万床；"
    f"教大有效需求约{EDUHK['eff']/10000:.2f}万床、校内供给约{EDUHK['beds']/10000:.2f}万床、"
    f"校外缺口约{EDUHK['gap']/10000:.2f}万床。"
)

RED = Font(name="Microsoft YaHei", size=11, color="C00000", bold=True)
BODY = Font(name="Microsoft YaHei", size=11)
HDR_FONT = Font(name="Microsoft YaHei", size=11, color="FFFFFF", bold=True)
HDR_FILL = PatternFill("solid", fgColor="C00000")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_cell(ws, r, c, value, font=BODY, fill=None, align=LEFT):
    cell = ws.cell(r, c, value)
    cell.font = font
    cell.border = BORDER
    cell.alignment = align
    if fill:
        cell.fill = fill
    return cell


def add_level_rows(ws, start_row: int, school: str, level: str, data: dict, m: dict) -> int:
    r = start_row
    local_nl = f"本地{data['local']:,}/非本地{data['nl']:,}"
    rows = [
        ("在校生（2024/25）", f"{data['total']:,}", local_nl),
        ("有效住宿需求", f"{m['eff']:,}", "本地全日制×35%＋非本地×100%"),
        ("校内宿位", f"~{data['beds']:,}" if data["beds"] else "共用", data["beds_note"]),
        ("校内覆盖率", m["cov"], f"{data['beds']:,}/{data['total']:,}" if data["beds"] else "—"),
        ("校外缺口", f"~{m['gap']:,}", "有效需求−校内宿位"),
    ]
    for i, (ind, num, note) in enumerate(rows):
        rr = r + i
        style_cell(ws, rr, 1, school if i == 0 else "", BODY, align=CENTER)
        style_cell(ws, rr, 2, level if i == 0 else "", BODY, align=CENTER)
        style_cell(ws, rr, 3, ind, BODY)
        style_cell(ws, rr, 4, num, RED if ind == "校外缺口" else BODY, align=CENTER)
        style_cell(ws, rr, 5, note, BODY)
    return r + len(rows)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "校外住宿缺口测算"

    ws.merge_cells("A1:E1")
    c1 = ws["A1"]
    c1.value = f"▶ {SENTENCE_1}"
    c1.font = BODY
    c1.alignment = WRAP

    ws.merge_cells("A2:E2")
    c2 = ws["A2"]
    c2.value = f"▶ {SENTENCE_2}"
    c2.font = BODY
    c2.alignment = WRAP

    headers = ["学校", "年级", "指标", "数字", "备注"]
    hr = 4
    for col, h in enumerate(headers, 1):
        style_cell(ws, hr, col, h, HDR_FONT, HDR_FILL, CENTER)

    r = 5
    r = add_level_rows(ws, r, "香港中文大学", "本科生", CUHK_UG, CUHK["ug"])
    r = add_level_rows(ws, r, "", "研究生", CUHK_PG, CUHK["pg"])
    style_cell(ws, r, 1, "", BODY, align=CENTER)
    style_cell(ws, r, 2, "合计缺口", BODY, align=CENTER)
    style_cell(ws, r, 3, "—", BODY, align=CENTER)
    style_cell(ws, r, 4, f"~{CUHK['gap']:,}", RED, align=CENTER)
    style_cell(ws, r, 5, "本科+研究生校外缺口合计", BODY)
    r += 1

    r = add_level_rows(ws, r, "香港教育大学", "本科生", EDUHK_UG, EDUHK["ug"])
    r = add_level_rows(ws, r, "", "研究生", EDUHK_PG, EDUHK["pg"])
    style_cell(ws, r, 1, "", BODY, align=CENTER)
    style_cell(ws, r, 2, "合计缺口", BODY, align=CENTER)
    style_cell(ws, r, 3, "—", BODY, align=CENTER)
    style_cell(ws, r, 4, f"~{EDUHK['gap']:,}", RED, align=CENTER)
    style_cell(ws, r, 5, "本科+研究生校外缺口合计", BODY)

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 36
    widths = [14, 10, 18, 14, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A{r+2}:E{r+2}")
    foot = ws.cell(r + 2, 1)
    foot.value = (
        "数据来源：CUHK Facts 2024（2024/9/30）；教大立法会 CB(3)315/2025 附件 Table 2（2024/10/31，FTE）；"
        "宿位：中大招生册、教大舍堂生活页。测算公式与高力国际全港学生住宿分析一致。"
    )
    foot.font = Font(name="Microsoft YaHei", size=9, color="666666")
    foot.alignment = WRAP

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(SENTENCE_1)
    print(SENTENCE_2)


if __name__ == "__main__":
    main()
