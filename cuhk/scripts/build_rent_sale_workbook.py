# -*- coding: utf-8 -*-
"""Build dual-sheet Excel: Apr/May 2026 rent + secondary sales for 13 estates."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "周邊樓盤租賃及成交對標表.xlsx"

ESTATES = [
    {
        "seq": 1,
        "name": "大埔寶馬山",
        "en": "Pacific Palisades",
        "dev": "信和",
        "loc": "山賢路8號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E5%A4%A7%E5%9F%94%E5%AF%B6%E9%A6%AC%E5%B1%B1-E11504",
        "rent": {
            "apr_rent": "18,000 – 30,000",
            "apr_psf": "32",
            "may_rent": "18,000 – 30,000",
            "may_psf": "32",
            "listings": "—",
            "avg_psf": "32",
            "dtype": "屋苑放盤參考",
            "note": "4–5月無租務成交；美聯屋苑專頁放盤租價區間",
        },
        "sale": {
            "apr_count": 4,
            "apr_psf_rng": "10,038 – 11,376",
            "apr_psf_avg": 10655,
            "may_count": 2,
            "may_psf_rng": "10,442 – 10,938",
            "may_psf_avg": 10628,
            "note": "4月含1宗^特殊成交",
        },
    },
    {
        "seq": 2,
        "name": "悠然山莊",
        "en": "The Paragon",
        "dev": "加文",
        "loc": "山賢路9號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E6%82%A0%E7%84%B6%E5%B1%B1%E8%8E%8A-E11584",
        "rent": {
            "apr_rent": "36,000",
            "apr_psf": "29",
            "may_rent": "36,000",
            "may_psf": "29",
            "listings": "1",
            "avg_psf": "29",
            "dtype": "屋苑放盤參考",
            "note": "4–5月無租務成交；僅1個在租放盤",
        },
        "sale": {
            "apr_count": 2,
            "apr_psf_rng": "9,069 – 9,497",
            "apr_psf_avg": 9283,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "5月無註冊處成交",
        },
    },
    {
        "seq": 3,
        "name": "天鑽",
        "en": "The Regent",
        "dev": "新鴻基",
        "loc": "創新路1號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E5%A4%A9%E9%91%BD-E000016716",
        "rent": {
            "apr_rent": "20,000 – 25,500",
            "apr_psf": "34 – 39",
            "may_rent": "15,000 – 31,000",
            "may_psf": "34 – 53",
            "listings": "48",
            "avg_psf": "41",
            "dtype": "租務成交＋放租叫價",
            "note": "4月：1宗租務成交@$38＋2盤放租叫價；5月：4宗租務成交＋當月美聯放租盤",
        },
        "sale": {
            "apr_count": 6,
            "apr_psf_rng": "12,533 – 15,011",
            "apr_psf_avg": 13518,
            "may_count": 6,
            "may_psf_rng": "12,490 – 13,721",
            "may_psf_avg": 13033,
            "note": "5月已剔除連車位特殊成交",
        },
    },
    {
        "seq": 4,
        "name": "雍怡雅苑",
        "en": "Chateau Royale",
        "dev": "新鴻基",
        "loc": "雍宜路1號",
        "midland_url": "https://www.midland.com.hk/zh-hk/property/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E9%9B%8D%E6%80%A1%E9%9B%85%E8%8B%91%E9%9B%8D%E6%97%A5%E5%BA%AD%E9%9B%8D%E5%AE%9C%E8%B7%AF1%E8%99%9F-%E7%8D%A8%E7%AB%8B%E5%B1%8B-M300300767",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "—",
            "avg_psf": "—",
            "dtype": "—",
            "note": "低密度獨立屋；美聯無獨立屋苑租務專頁及4–5月租務紀錄",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "最近成交2025年6月",
        },
    },
    {
        "seq": 5,
        "name": "偉東·雍宜山莊",
        "en": "Grand Dynasty View",
        "dev": "—",
        "loc": "下黃宜坳",
        "midland_url": "",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "—",
            "avg_psf": "—",
            "dtype": "—",
            "note": "美聯未設屋苑專頁",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "美聯未設屋苑專頁",
        },
    },
    {
        "seq": 6,
        "name": "翡翠花園",
        "en": "Savanna Garden",
        "dev": "新鴻基",
        "loc": "大埔公路大埔滘段4283號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E7%BF%A1%E7%BF%A0%E8%8A%B1%E5%9C%92-E00070",
        "rent": {
            "apr_rent": "28,500 – 40,000",
            "apr_psf": "33",
            "may_rent": "28,500 – 40,000",
            "may_psf": "33",
            "listings": "—",
            "avg_psf": "33",
            "dtype": "屋苑放盤參考",
            "note": "4–5月無租務成交；美聯屋苑專頁放盤租價區間",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "最近成交2026年3月",
        },
    },
    {
        "seq": 7,
        "name": "疊翠豪庭",
        "en": "The Paramount (4188)",
        "dev": "—",
        "loc": "大埔公路大埔滘段4188號",
        "midland_url": "",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "—",
            "avg_psf": "—",
            "dtype": "—",
            "note": "與盈峰翠邸（山塘路23號）為不同項目；美聯未設屋苑專頁",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "美聯未設屋苑專頁",
        },
    },
    {
        "seq": 8,
        "name": "泓山",
        "en": "Hampstead",
        "dev": "建灝",
        "loc": "逸遙路3號",
        "midland_url": "",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "—",
            "avg_psf": "—",
            "dtype": "—",
            "note": "美聯未設屋苑專頁（勿與九肚山駿嶺薈混淆）",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "美聯未設屋苑專頁",
        },
    },
    {
        "seq": 9,
        "name": "逍遙雋岸",
        "en": "L Utopie",
        "dev": "興聯置業",
        "loc": "逸遙路大埔滘段18號",
        "midland_url": "https://www.midland.com.hk/zh-hk/estate/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E9%80%8D%E9%81%99%E9%9B%8B%E5%B2%B8-E000004058",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "0",
            "avg_psf": "—",
            "dtype": "—",
            "note": "美聯屋苑專頁無在租放盤及4–5月租務紀錄",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "最近成交2017年",
        },
    },
    {
        "seq": 10,
        "name": "峰林",
        "en": "The Peak",
        "dev": "—",
        "loc": "大埔滘",
        "midland_url": "",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "—",
            "avg_psf": "—",
            "dtype": "—",
            "note": "美聯未設屋苑專頁",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "美聯未設屋苑專頁",
        },
    },
    {
        "seq": 11,
        "name": "史提福樓",
        "en": "Stafford House",
        "dev": "—",
        "loc": "大埔滘新圍",
        "midland_url": "",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "—",
            "avg_psf": "—",
            "dtype": "—",
            "note": "美聯未設屋苑專頁",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "美聯未設屋苑專頁",
        },
    },
    {
        "seq": 12,
        "name": "僑東·羅宜山莊",
        "en": "Kiu Tung Lo Yee Villa",
        "dev": "—",
        "loc": "大埔滘",
        "midland_url": "",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "—",
            "avg_psf": "—",
            "dtype": "—",
            "note": "美聯未設屋苑專頁",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 0,
            "may_psf_rng": "—",
            "may_psf_avg": None,
            "note": "美聯未設屋苑專頁",
        },
    },
    {
        "seq": 13,
        "name": "盈峰翠邸",
        "en": "The Paramount",
        "dev": "長江實業",
        "loc": "山塘路23號",
        "midland_url": "https://www.midland.com.hk/zh-hk/property/%E6%96%B0%E7%95%8C-%E5%A4%A7%E5%9F%94%E5%8D%8A%E5%B1%B1-%E7%9B%88%E5%B3%B0%E7%BF%A0%E9%82%B802%E5%BA%A7%E4%B8%AD%E5%B1%A4B%E5%AE%A4-M300551705",
        "rent": {
            "apr_rent": "—",
            "apr_psf": "—",
            "may_rent": "—",
            "may_psf": "—",
            "listings": "—",
            "avg_psf": "—",
            "dtype": "—",
            "note": "美聯無獨立屋苑租務專頁；4–5月無租務成交紀錄",
        },
        "sale": {
            "apr_count": 0,
            "apr_psf_rng": "—",
            "apr_psf_avg": None,
            "may_count": 1,
            "may_psf_rng": "9,279",
            "may_psf_avg": 9279,
            "note": "5月28日2座2樓B室916呎@$9,279",
        },
    },
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
META_FILL = PatternFill("solid", fgColor="D9E2F3")


def style_header_row(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 0
        for cell in col:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = max(width, 10)


def write_meta(ws, lines):
    for i, line in enumerate(lines, 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)
        cell = ws.cell(row=i, column=1, value=line)
        cell.fill = META_FILL
        cell.alignment = Alignment(wrap_text=True)


def build_rent_sheet(wb):
    ws = wb.active
    ws.title = "租賃對標"
    write_meta(
        ws,
        [
            "大埔黃宜坳周邊樓盤 · 2026年4–5月租賃對標（13個可比屋苑）",
            "資料來源：美聯物業；優先採當月租務成交，其次當月美聯放租叫價，最後屋苑放盤參考區間",
            "整理日期：2026-06-10｜月租單位：港元/月｜呎租單位：港元/平方呎（實用）",
        ],
    )
    headers = [
        "序號",
        "項目名稱",
        "英文名",
        "發展商",
        "位置",
        "2026年4月月租範圍（元）",
        "2026年4月呎租範圍（元/呎）",
        "2026年5月月租範圍（元）",
        "2026年5月呎租範圍（元/呎）",
        "美聯在租盤數",
        "放盤平均呎租（元/呎）",
        "資料類型",
        "備註",
        "美聯資料連結",
    ]
    start = 5
    for j, h in enumerate(headers, 1):
        ws.cell(row=start, column=j, value=h)
    style_header_row(ws, start, len(headers))

    for i, e in enumerate(ESTATES, start + 1):
        r = e["rent"]
        ws.append(
            [
                e["seq"],
                e["name"],
                e["en"],
                e["dev"],
                e["loc"],
                r["apr_rent"],
                r["apr_psf"],
                r["may_rent"],
                r["may_psf"],
                r["listings"],
                r["avg_psf"],
                r["dtype"],
                r["note"],
                e["midland_url"] or "—",
            ]
        )
    autosize(ws)


def build_sale_sheet(wb):
    ws = wb.create_sheet("二手成交")
    write_meta(
        ws,
        [
            "大埔黃宜坳周邊樓盤 · 2026年4–5月二手成交對標（13個可比屋苑）",
            "資料來源：美聯物業屋苑專頁／放盤頁註冊處成交紀錄；無成交標示為0",
            "平均呎價算法：面積加權平均（∑成交價÷∑實用面積）｜整理日期：2026-06-10",
        ],
    )
    headers = [
        "序號",
        "項目名稱",
        "英文名",
        "發展商",
        "位置",
        "2026年4月成交量（宗）",
        "2026年4月呎價範圍（元/呎）",
        "2026年4月平均呎價（元/呎）",
        "2026年5月成交量（宗）",
        "2026年5月呎價範圍（元/呎）",
        "2026年5月平均呎價（元/呎）",
        "備註",
        "美聯資料連結",
    ]
    start = 5
    for j, h in enumerate(headers, 1):
        ws.cell(row=start, column=j, value=h)
    style_header_row(ws, start, len(headers))

    for e in ESTATES:
        s = e["sale"]
        ws.append(
            [
                e["seq"],
                e["name"],
                e["en"],
                e["dev"],
                e["loc"],
                s["apr_count"],
                s["apr_psf_rng"],
                s["apr_psf_avg"] if s["apr_psf_avg"] is not None else "—",
                s["may_count"],
                s["may_psf_rng"],
                s["may_psf_avg"] if s["may_psf_avg"] is not None else "—",
                s["note"],
                e["midland_url"] or "—",
            ]
        )
    autosize(ws)


def main():
    wb = Workbook()
    build_rent_sheet(wb)
    build_sale_sheet(wb)
    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
