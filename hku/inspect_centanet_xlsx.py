# -*- coding: utf-8 -*-
import json
import openpyxl
from pathlib import Path

p = Path(__file__).parent / "港大周邊私人住宅" / "周邊項目租成交情況.xlsx"
wb = openpyxl.load_workbook(p)
ws = wb.active
rows = []
for r in range(1, ws.max_row + 1):
    rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
def default(o):
    return o.isoformat() if hasattr(o, "isoformat") else str(o)

Path(__file__).parent.joinpath("centanet_xlsx_structure.json").write_text(
    json.dumps(rows, ensure_ascii=False, indent=2, default=default), encoding="utf-8"
)
print("wrote", len(rows), "rows")
