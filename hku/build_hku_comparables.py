# -*- coding: utf-8 -*-
"""Build HKU comparables rent table from Midland data."""
import json
import re
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
CUTOFF_YEAR = datetime.now().year - 15

# name, url, page_type (estate|new)
ESTATES = [
    ("吉喆", "https://www.midland.com.hk/zh-hk/new-property/%E6%B8%AF%E5%B3%B6-%E5%A0%85%E5%B0%BC%E5%9C%B0%E5%9F%8E-%E5%90%89%E5%96%86-E000019385", "new"),
    ("Eight South Lane", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E5%A0%85%E5%B0%BC%E5%9C%B0%E5%9F%8E-Eight-South-Lane-E000015221", "estate"),
    ("藝里坊．1號", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E8%97%9D%E9%87%8C%E5%9D%8A%EF%BC%8E1%E8%99%9F-E000016617", "estate"),
    ("翰林峰", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E7%BF%B0%E6%9E%97%E5%B3%B0-E000015984", "estate"),
    ("曉譽", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E6%9B%89%E8%AD%BD-E000014006", "estate"),
    ("尚瓏", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E5%B0%9A%E7%93%8F-E000017130", "estate"),
    ("懿山", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E6%87%BF%E5%B1%B1-E000013810", "estate"),
    ("瑧蓺", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E7%91%A7%E8%93%BA-E000016134", "estate"),
    ("高士台", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-Kensington-Hill-E000015321", "estate"),
    ("63 Pokfulam", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E6%B8%AF%E5%B3%B6%E8%A5%BF-63-Pokfulam-E000016034", "estate"),
    ("Kennedy 38", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-Kennedy-38-E000016716", "estate"),
    ("藝里坊．2號", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E8%97%9D%E9%87%8C%E5%9D%8A-2%E8%99%9F-E000017106", "estate"),
    ("赋居", "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-Resiglow-Bonham-E000016875", "estate"),
]

ROOM_MAP = {"0": "studio", "1": "one", "2": "two"}

# 吉喆：美聯新盤頁無放租統計；開放式取自美聯新聞租務成交（211呎@約$17,300）
JIZHE_MANUAL = {
    "year": 2025,
    "studio": {"area": "206–297", "rent": "16,500–18,500", "psf": "78–85"},
}

# 高士台：僅2房以上，美聯屋苑頁放盤租價（2026年6月）
KENSINGTON_MANUAL = {
    "two": {"area": "532–864", "rent": "38,000–57,000", "psf": "54–62"},
}

# 赋居(Resiglow-Bonham)：美聯註冊租務成交參考（2024-02）
BONHAM_MANUAL = {
    "one": {"area": "312", "rent": "22,800", "psf": "73"},
}


def fetch(url):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=40) as r:
        return r.read().decode("utf-8", "replace")


def parse_next_data(html):
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    return json.loads(m.group(1)) if m else None


def get_estate_dict(page_props, page_type):
    if page_type == "new":
        return page_props.get("newPropertyDetail") or {}
    result = page_props.get("result") or {}
    return result.get("estateData") or {}


def extract_year(ed):
    for key in ("first_op_date", "last_op_date", "op_date"):
        val = ed.get(key)
        if isinstance(val, str):
            m = re.search(r"(\d{4})", val)
            if m:
                return int(m.group(1))
    return None


def fmt_range(lo, hi, suffix=""):
    if lo is None and hi is None:
        return None
    lo = lo if lo is not None else hi
    hi = hi if hi is not None else lo
    if lo == hi:
        return f"{lo:,}{suffix}"
    return f"{lo:,}–{hi:,}{suffix}"


def area_from_rent(rent_lo, rent_hi, psf_lo, psf_hi):
    areas = []
    if rent_lo and psf_hi:
        areas.append(round(rent_lo / psf_hi))
    if rent_hi and psf_lo:
        areas.append(round(rent_hi / psf_lo))
    if not areas:
        return None
    return fmt_range(min(areas), max(areas))


def merge_area(geo_areas, calc_area, estate_min, estate_max):
    areas = list(geo_areas)
    if calc_area:
        for part in re.findall(r"[\d,]+", calc_area):
            areas.append(int(part.replace(",", "")))
    if not areas and estate_min and estate_max:
        return fmt_range(estate_min, estate_max)
    if not areas:
        return None
    return fmt_range(min(areas), max(areas))


def room_stats(ed):
    prop_stat = ed.get("property_stat") or {}
    rooms = prop_stat.get("rooms") or {}
    geo_rent = (ed.get("geo_transaction") or {}).get("secondhand_rent") or []
    geo_by_bed = {}
    for tx in geo_rent:
        b = str(tx.get("bedroom", ""))
        if tx.get("net_area"):
            geo_by_bed.setdefault(b, []).append(tx["net_area"])
    return rooms, geo_by_bed


def build_room_info(bed_key, rooms, geo_by_bed, estate_min, estate_max):
    rinfo = rooms.get(bed_key) or {}
    rent_lo = rinfo.get("min_rent")
    rent_hi = rinfo.get("max_rent")
    psf_lo = rinfo.get("min_net_ft_rent")
    psf_hi = rinfo.get("max_net_ft_rent")
    rent_count = rinfo.get("rent_count", 0)

    if not rent_count and not rent_lo:
        return None

    calc_area = area_from_rent(rent_lo, rent_hi, psf_lo, psf_hi)
    geo_areas = geo_by_bed.get(bed_key, [])
    return {
        "area": merge_area(geo_areas, calc_area, estate_min, estate_max),
        "rent": fmt_range(rent_lo, rent_hi),
        "psf": fmt_range(psf_lo, psf_hi),
        "rent_count": rent_count,
    }


def parse_estate(name, url, page_type):
    html = fetch(url)
    if "屋苑搜尋" in html and "共找到" in html:
        return {"name": name, "url": url, "error": "invalid url"}

    data = parse_next_data(html)
    if not data:
        return {"name": name, "url": url, "error": "no __NEXT_DATA__"}

    page_props = data["props"]["pageProps"]
    ed = get_estate_dict(page_props, page_type)
    if not ed:
        return {"name": name, "url": url, "error": "no estate data"}

    row = {
        "name": name,
        "url": url,
        "year": extract_year(ed),
        "source": f"美聯：{url}",
    }

    if name == "吉喆":
        row["year"] = JIZHE_MANUAL["year"]
        for col in ["studio", "one", "two"]:
            row[col] = JIZHE_MANUAL.get(col)
        row["source"] += "；開放式租務參考：美聯新聞（2025年租務成交）"
        return row

    rooms, geo_by_bed = room_stats(ed)
    emin, emax = ed.get("min_net_area"), ed.get("max_net_area")

    for bed_key, col in ROOM_MAP.items():
        row[col] = build_room_info(bed_key, rooms, geo_by_bed, emin, emax)

    if name == "高士台":
        row["two"] = KENSINGTON_MANUAL.get("two")
        row["source"] += "；兩房租金：美聯屋苑頁放盤叫價"
    if name == "赋居":
        row["one"] = BONHAM_MANUAL.get("one")
        row["source"] += "；一房租金：美聯註冊租務成交（2024-02）"

    return row


def write_xlsx(rows, xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for r in range(4, ws.max_row + 1):
        for c in range(1, 13):
            ws.cell(r, c).value = None

    for i, row in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1).value = row["name"]
        ws.cell(r, 2).value = row.get("year")
        for j, col in enumerate(["studio", "one", "two"]):
            base = 3 + j * 3
            info = row.get(col)
            if info:
                ws.cell(r, base).value = info.get("area")
                ws.cell(r, base + 1).value = info.get("rent")
                ws.cell(r, base + 2).value = info.get("psf")
        ws.cell(r, 12).value = row.get("source")

    wb.save(xlsx_path)


def main():
    results = []
    for name, url, ptype in ESTATES:
        try:
            row = parse_estate(name, url, ptype)
            if row.get("error"):
                print(f"ERROR {name}: {row['error']}")
                continue
            if row.get("year") and row["year"] < CUTOFF_YEAR:
                print(f"SKIP {name}: year {row['year']}")
                continue
            results.append(row)
            print(f"OK {name} ({row.get('year')})")
        except Exception as e:
            print(f"FAIL {name}: {e}")

    out_json = Path(__file__).parent / "hku_rent_data.json"
    out_json.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    xlsx = Path(__file__).parent / "港大周邊私人住宅" / "HKU comparables.xlsx"
    write_xlsx(results, xlsx)
    print(f"Wrote {len(results)} rows -> {xlsx}")


if __name__ == "__main__":
    main()
