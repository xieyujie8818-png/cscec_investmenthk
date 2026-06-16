# -*- coding: utf-8 -*-
"""Update HKU comparables 成交.xlsx with Midland rent transactions (90 days).

Uses Midland transaction list pages with 租樓 + 90天內 filters, matching the
site's own search UI (not estate-page transactionDataObj).
"""
from __future__ import annotations

import json
import re
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
AS_OF = date.today()
CUTOFF = AS_OF - timedelta(days=90)

XLSX = Path(__file__).parent / "港大周邊私人住宅" / "HKU comparables 成交.xlsx"
BASE = "https://www.midland.com.hk"

ESTATE_PAGES = {
    "吉喆": "https://www.midland.com.hk/zh-hk/new-property/%E6%B8%AF%E5%B3%B6-%E5%A0%85%E5%B0%BC%E5%9C%B0%E5%9F%8E-%E5%90%89%E5%96%86-E000019385",
    "Eight South Lane": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E5%A0%85%E5%B0%BC%E5%9C%B0%E5%9F%8E-Eight-South-Lane-E000015221",
    "藝里坊．1號": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E8%97%9D%E9%87%8C%E5%9D%8A%EF%BC%8E1%E8%99%9F-E000016617",
    "翰林峰": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E7%BF%B0%E6%9E%97%E5%B3%B0-E000015984",
    "曉譽": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E6%9B%89%E8%AD%BD-E000014006",
    "尚瓏": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E5%B0%9A%E7%93%8F-E000017130",
    "懿山": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E6%87%BF%E5%B1%B1-E000013810",
    "瑧蓺": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E7%91%A7%E8%93%BA-E000016134",
    "高士台": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-Kensington-Hill-E000015321",
    "63 Pokfulam": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E6%B8%AF%E5%B3%B6%E8%A5%BF-63-Pokfulam-E000016034",
    "Kennedy 38": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-Kennedy-38-E000016716",
    "藝里坊．2號": "https://www.midland.com.hk/zh-hk/estate/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E8%97%9D%E9%87%8C%E5%9D%8A-2%E8%99%9F-E000017106",
    "尚逸": "https://www.midland.com.hk/zh-hk/new-property/%E6%B8%AF%E5%B3%B6-%E8%A5%BF%E7%87%9F%E7%9B%A4-%E5%B0%9A%E9%80%B8-E000019401",
}

BED_MAP = {"0": "studio", "1": "one", "2": "two"}

APPLY_FILTERS_JS = """
async () => {
  const sleep = ms => new Promise(r => setTimeout(r, ms));
  const click = prefix => {
    const el = [...document.querySelectorAll('label')].find(
      l => l.textContent.includes(prefix)
    );
    if (el) { el.click(); return true; }
    return false;
  };
  click('租樓 (');
  await sleep(1500);
  click('90天內 (');
  await sleep(2500);
}
"""

COLLECT_TX_JS = """
() => {
  const links = [...document.querySelectorAll('a')].filter(a => {
    const t = a.textContent || '';
    return /\\d{2}\\/\\d{2}\\/\\d{4}/.test(t) && !t.includes('萬');
  });
  return [...new Set(links.map(a => a.textContent.replace(/\\s+/g, ' ').trim()))];
}
"""


def fetch_html(url: str) -> str:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=45) as r:
        return r.read().decode("utf-8", "replace")


def est_id_from_url(url: str) -> str:
    m = re.search(r"(E\d{9})", url)
    if not m:
        raise ValueError(f"No estate id in {url}")
    return m.group(1)


def tx_list_path(estate_page_url: str) -> str:
    est_id = est_id_from_url(estate_page_url)
    html = fetch_html(estate_page_url)
    paths = re.findall(r"/zh-hk/list/transaction/[^\"']+", html)
    own = [p for p in paths if est_id in p]
    if not own:
        raise ValueError(f"No transaction list link for {est_id}")
    return own[0]


def parse_op_date(html: str) -> int | None:
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if not m:
        return None
    pp = json.loads(m.group(1))["props"]["pageProps"]
    raw = None
    if "newPropertyDetail" in pp:
        raw = (pp.get("newPropertyDetail") or {}).get("first_op_date")
    else:
        raw = ((pp.get("result") or {}).get("estateData") or {}).get("first_op_date")
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).year
    except ValueError:
        return None


def parse_tx_date(s: str) -> date | None:
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
    if not m:
        return None
    return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))


def parse_rent_card(text: str) -> dict | None:
    if "萬" in text:
        return None
    dm = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    if not dm:
        return None
    d = parse_tx_date(dm.group(1))
    if d is None or d < CUTOFF or d > AS_OF:
        return None

    if "開放式" in text:
        bed = "0"
    elif re.search(r"1房", text):
        bed = "1"
    elif re.search(r"2房", text):
        bed = "2"
    else:
        return None

    rent = area = psf = None
    rent_m = (
        re.search(r"租\$(\d{1,3}(?:,\d{3})*)", text)
        or re.search(r"房\$(\d{1,3}(?:,\d{3})*)", text)
        or re.search(r"開放式\$(\d{1,3}(?:,\d{3})*)", text)
    )
    if rent_m:
        rent = int(rent_m.group(1).replace(",", ""))

    area_m = re.search(r"實(\d{2,4})呎", text)
    if area_m:
        area = int(area_m.group(1))
    else:
        glued = re.search(r"\$(\d{1,3}(?:,\d{3})*)(\d{2,4})呎", text)
        if glued:
            if rent is None:
                rent = int(glued.group(1).replace(",", ""))
            area = int(glued.group(2))

    psf_m = re.search(r"@\$(\d+)", text) or re.search(r"呎\$(\d+)", text)
    if psf_m:
        psf = int(psf_m.group(1))

    if rent is None or area is None or psf is None:
        return None
    if rent < 5000 or psf >= 200 or area < 100 or area > 2000:
        return None

    return {"date": d, "bedroom": bed, "area": area, "rent": rent, "psf": psf}


def fmt_range(lo, hi):
    if lo is None and hi is None:
        return None
    lo = lo if lo is not None else hi
    hi = hi if hi is not None else lo
    if lo == hi:
        return f"{lo:,}"
    return f"{lo:,}–{hi:,}"


def summarize_group(txs: list[dict]) -> dict | None:
    if not txs:
        return None
    areas = [t["area"] for t in txs]
    rents = [t["rent"] for t in txs]
    psfs = [t["psf"] for t in txs]
    return {
        "area": fmt_range(min(areas), max(areas)),
        "rent": fmt_range(min(rents), max(rents)),
        "psf": fmt_range(min(psfs), max(psfs)),
        "count": len(txs),
    }


def collect_all_pages(page) -> list[str]:
    seen: set[str] = set()
    while True:
        cards = page.evaluate(COLLECT_TX_JS)
        seen.update(cards)
        next_btn = page.locator('nav[aria-label="Pagination"] a', has_text=re.compile(r"^2$|^下一頁$"))
        if next_btn.count() == 0:
            break
        # Only paginate when a page-2 link exists and we haven't clicked it
        page2 = page.locator('nav[aria-label="Pagination"] a', has_text="2")
        if page2.count() == 0:
            break
        if page.locator('nav[aria-label="Pagination"] a[aria-current="page"]', has_text="2").count():
            break
        page2.first.click()
        page.wait_for_timeout(2500)
    return list(seen)


def scrape_estate_rent_tx(page, estate_page_url: str) -> tuple[list[dict], str]:
    path = tx_list_path(estate_page_url)
    tx_url = f"{BASE}{path}?tx_type=L"
    page.goto(tx_url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(3000)
    page.evaluate(APPLY_FILTERS_JS)
    page.wait_for_timeout(2000)

    raw_cards = collect_all_pages(page)
    txs = []
    for card in raw_cards:
        tx = parse_rent_card(card)
        if tx:
            txs.append(tx)

    # dedupe
    seen = set()
    deduped = []
    for t in txs:
        key = (t["date"], t["bedroom"], t["area"], t["rent"], t["psf"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(t)

    return deduped, tx_url


def build_row(name: str, year_hint, txs: list[dict], source_url: str) -> dict:
    by_bed = {k: [] for k in BED_MAP}
    for t in txs:
        bed = t["bedroom"]
        if bed in by_bed:
            by_bed[bed].append(t)

    row = {"name": name, "year": year_hint}
    for bed, col in BED_MAP.items():
        row[col] = summarize_group(by_bed[bed])
    row["source"] = f"美聯租務成交（近90日）：{source_url}"
    return row


def read_estates_from_xlsx() -> list[tuple[str, int | None]]:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    estates = []
    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not isinstance(name, str):
            continue
        if name.startswith("数据说明") or name.startswith("數據說明"):
            continue
        year = ws.cell(r, 2).value
        estates.append((name.strip(), year))
    return estates


def write_xlsx(rows: list[dict], note_row: int):
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    for i, row in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1).value = row["name"]
        if row.get("year"):
            ws.cell(r, 2).value = row["year"]
        for j, col in enumerate(["studio", "one", "two"]):
            base = 3 + j * 3
            info = row.get(col)
            if info:
                ws.cell(r, base).value = info.get("area")
                ws.cell(r, base + 1).value = info.get("rent")
                ws.cell(r, base + 2).value = info.get("psf")
            else:
                ws.cell(r, base).value = None
                ws.cell(r, base + 1).value = None
                ws.cell(r, base + 2).value = None

    note = (
        f"数据说明\n"
        f"租金口径：美联租务成交（注册处／美联记录），近90日内（{CUTOFF.isoformat()} 至 {AS_OF.isoformat()}）。"
        f"按开放式（0房）、一房、两房分别汇总面积、月租及呎租区间。"
        f"数据来源：美联成交搜索页（租楼＋90天内筛选）。"
    )
    ws.cell(note_row, 1).value = note
    wb.save(XLSX)


def main():
    estates = read_estates_from_xlsx()
    note_row = 4 + len(estates)
    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.set_extra_http_headers({"Accept-Language": "zh-HK,zh;q=0.9"})

        for name, year_hint in estates:
            estate_url = ESTATE_PAGES.get(name)
            if not estate_url:
                print(f"SKIP {name}: no URL mapping")
                results.append({"name": name, "year": year_hint, "studio": None, "one": None, "two": None})
                continue
            try:
                year = parse_op_date(fetch_html(estate_url)) or year_hint
                txs, tx_url = scrape_estate_rent_tx(page, estate_url)
                row = build_row(name, year, txs, tx_url)
                results.append(row)
                print(f"OK {name}: {len(txs)} rent tx in 90d")
                for col in ["studio", "one", "two"]:
                    if row.get(col):
                        print(f"  {col}: {row[col]}")
            except Exception as e:
                print(f"FAIL {name}: {e}")
                results.append({"name": name, "year": year_hint, "studio": None, "one": None, "two": None})

        browser.close()

    out = Path(__file__).parent / "hku_rent_tx_90d.json"
    out.write_text(json.dumps(results, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_xlsx(results, note_row)
    print(f"Wrote {XLSX}")


if __name__ == "__main__":
    main()
