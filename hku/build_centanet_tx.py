# -*- coding: utf-8 -*-
"""Fill 周邊項目租成交情況.xlsx from Centaline rent transactions (3 years, top 3 each)."""
from __future__ import annotations

import json
import re
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
AS_OF = date.today()
CUTOFF = AS_OF - timedelta(days=3 * 365)
BASE = "https://hk.centanet.com/findproperty/zh-cn"

XLSX = Path(__file__).parent / "港大周邊私人住宅" / "周邊項目租成交情況.xlsx"
OUT_JSON = Path(__file__).parent / "centanet_rent_tx.json"

PROJECTS = [
    ("吉喆", "%E5%90%89%E5%96%86_2-SSSPWWPSWS", ""),
    ("翰林峰", "%E7%BF%B0%E6%9E%97%E5%B3%B0_2-SSPPWWPOWG", ""),
    ("Kennedy 38", "KENNEDY%2038_2-SSSPWWPVWS", ""),
    ("嘉林閣", "%E5%98%89%E6%9E%97%E9%96%A3_2-SDBBPPKJPS", ""),
    ("碧瑤灣", "%E7%A2%A7%E7%91%B6%E6%B9%BE_2-OJQCFRCORO", "?q=udjgk03it5jzb"),
    ("豪峰II", "Royalton-II_2-SDPPWWPJWS", ""),
    ("美林園（又名美琳苑）", "%E7%BE%8E%E7%90%B3%E5%9C%92_1-TITNZHTXHT", ""),
    ("Victoria Coast", "-Victoria%20Coast_2-SDPPWWPOWS", ""),
]


def fetch(url: str) -> str:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=45) as r:
        return r.read().decode("utf-8", "replace")


def extract_rent_paths(html: str) -> list[str]:
    paths = set()
    for m in re.finditer(r"transaction-detail\\u002F([^\"\\]+)", html):
        paths.add(m.group(1))
    for m in re.finditer(r"/transaction-detail/([^\"'\\]+)", html):
        paths.add(m.group(1))
    rent = [p for p in paths if re.search(r"[A-Z]{3}\d{6}R\d+", p)]
    return sorted(rent, reverse=True)


def strip_tags(html: str) -> str:
    text = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.DOTALL | re.I)
    text = re.sub(r"<style[^>]*>.*?</style>", " ", text, flags=re.DOTALL | re.I)
    text = re.sub(r"<[^>]+>", "\n", text)
    return re.sub(r"\n+", "\n", text)


def pick(text: str, pattern: str) -> str | None:
    m = re.search(pattern, text, re.I | re.S)
    return m.group(1).strip() if m else None


def parse_detail(html: str) -> dict | None:
    if "已租" not in html and "已租" not in html:
        return None
    if "出租价" not in html and "出租價" not in html:
        return None

    text = strip_tags(html)
    title = pick(html, r"<h1[^>]*>([^<]+)")
    if not title:
        title = pick(html, r'line1:"([^"]+)"')
    if title:
        title = re.sub(r"\s*中原[集團集团]+.*", "", title).strip()

    tx_date = pick(text, r"成交日期[：:]\s*(\d{4}-\d{2}-\d{2})")
    rent = pick(text, r"出租[价價][\s\S]*?\$\s*([\d,]+)")
    if not rent:
        rent = pick(text, r"已租[\s\S]{0,200}?\$\s*([\d,]+)")
    area = pick(text, r"实用[\s\S]*?(\d[\d,]*)\s*呎") or pick(text, r"實用[\s\S]*?(\d[\d,]*)\s*呎")
    psf = pick(text, r"实用[\s\S]*?\$(\d+)\s*/\s*呎") or pick(text, r"實用[\s\S]*?\$(\d+)\s*/\s*呎")
    if area and not psf:
        psf = pick(text, r"(\d[\d,]*)\s*呎\s*\n\s*\$(\d+)\s*/\s*呎")
        if psf and isinstance(psf, str) and "," in psf:
            # wrong group - retry
            m = re.search(r"(\d[\d,]*)\s*呎\s*\n\s*\$(\d+)\s*/\s*呎", text)
            psf = m.group(2) if m else None

    layout = pick(text, r"间隔\s*([^\n]+)") or pick(text, r"間隔\s*([^\n]+)")
    unit_type = pick(html, r'unitType:\{[^}]*?value:"([^"]+)"')
    if unit_type and re.fullmatch(r"T\d+", unit_type):
        layout = unit_type
    age = pick(text, r"楼龄\s*(\d+)") or pick(text, r"樓齡\s*(\d+)")
    direction = pick(text, r"座向\s*([^\n]+)")

    rent_i = int(str(rent).replace(",", "")) if rent else None
    area_i = int(str(area).replace(",", "")) if area else None
    psf_i = int(psf) if psf else None
    if rent_i and area_i and not psf_i:
        psf_i = round(rent_i / area_i)

    return {
        "title": title,
        "tx_date": tx_date,
        "rent": rent_i,
        "area": area_i,
        "psf": psf_i,
        "layout": layout,
        "age": int(age) if age else None,
        "direction": direction,
    }


def parse_addr_parts(title: str | None) -> tuple[str | None, str | None, str | None]:
    if not title:
        return None, None, None
    t = title.strip()
    unit_no = None
    m = re.search(r"([A-Z]室|\d+室)", t)
    if m:
        unit_no = m.group(1)
        t = t.replace(unit_no, "").strip()
    floor = None
    for f in ["高層", "高层", "中層", "中层", "低層", "低层", "地下"]:
        if f in t:
            floor = f.replace("层", "層")
            t = t.replace(f, "").strip()
            break
    m2 = re.search(r"(\d+座|[A-Z]座)", t)
    block = m2.group(1) if m2 else None
    return block, floor, unit_no


def normalize_layout(layout: str | None) -> str | None:
    if not layout:
        return None
    layout = layout.strip()
    m = re.search(r"(T\d+|\d房|開放式|开放式)", layout)
    return m.group(1) if m else layout.split()[0][:12]


def parse_tx_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def detail_url(path: str) -> str:
    return f"{BASE}/transaction-detail/{urllib.parse.quote(path, safe='-_')}"


def scrape_project(name: str, slug: str, query: str) -> dict:
    list_url = f"{BASE}/list/transaction/{slug}{query}"
    html = fetch(list_url)
    paths = extract_rent_paths(html)

    records = []
    for path in paths:
        try:
            dhtml = fetch(detail_url(path))
            d = parse_detail(dhtml)
        except Exception:
            continue
        if not d or not d.get("rent") or not d.get("area"):
            continue
        tx_date = parse_tx_date(d.get("tx_date"))
        if tx_date is None or tx_date < CUTOFF or tx_date > AS_OF:
            continue
        block, floor, unit_no = parse_addr_parts(d.get("title"))
        psf = d.get("psf")
        records.append(
            {
                "date": tx_date.isoformat(),
                "block": block,
                "floor": floor,
                "unit_no": unit_no,
                "layout": normalize_layout(d.get("layout")),
                "area": d["area"],
                "age": d.get("age"),
                "rent": d["rent"],
                "psf": psf,
                "special_psf": round(psf * 1.1, 1) if psf else None,
                "direction": d.get("direction"),
                "url": detail_url(path),
                "title": d.get("title"),
            }
        )

    records.sort(key=lambda x: x["date"], reverse=True)
    # dedupe by date+rent+area
    seen = set()
    deduped = []
    for rec in records:
        key = (rec["date"], rec["rent"], rec["area"], rec.get("unit_no"))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(rec)
    records = deduped[:3]

    return {
        "name": name,
        "list_url": list_url,
        "records": records,
        "note": None if records else "三年內無租務成交（中原）",
    }


def read_project_rows() -> list[tuple[int, str]]:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    projects = []
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        pname = ws.cell(r, 2).value
        if seq and pname and isinstance(pname, str):
            projects.append((int(seq), pname.strip()))
    return projects


def write_xlsx(results: list[dict]):
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    start_rows: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        pname = ws.cell(r, 2).value
        if seq and pname:
            start_rows[str(pname).strip()] = r

    for res in results:
        pname = res["name"]
        r0 = start_rows.get(pname)
        if not r0:
            continue
        recs = res["records"]
        for off in range(3):
            r = r0 + off
            if off < len(recs):
                rec = recs[off]
                ws.cell(r, 3).value = datetime.strptime(rec["date"], "%Y-%m-%d")
                ws.cell(r, 4).value = rec.get("block")
                ws.cell(r, 5).value = rec.get("floor")
                ws.cell(r, 6).value = rec.get("unit_no")
                ws.cell(r, 7).value = rec.get("layout")
                ws.cell(r, 8).value = rec.get("area")
                ws.cell(r, 9).value = rec.get("age")
                ws.cell(r, 10).value = rec.get("rent")
                ws.cell(r, 11).value = rec.get("psf")
                ws.cell(r, 12).value = rec.get("special_psf")
            else:
                if off == 0 and not recs:
                    ws.cell(r, 3).value = res.get("note") or "三年內無租務成交（中原）"
                else:
                    ws.cell(r, 3).value = None
                for c in range(4, 13):
                    ws.cell(r, c).value = None

    wb.save(XLSX)


def main():
    project_map = {n: (s, q) for n, s, q in PROJECTS}
    results = []
    for _seq, name in read_project_rows():
        slug, query = project_map.get(name, ("", ""))
        if not slug:
            results.append({"name": name, "records": [], "note": "未配置中原搜尋連結"})
            continue
        try:
            res = scrape_project(name, slug, query)
            results.append(res)
        except Exception as e:
            results.append({"name": name, "records": [], "note": f"抓取失敗: {e}"})

    OUT_JSON.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx(results)


if __name__ == "__main__":
    main()
